from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests

# Função para obter os dados da tabela
def obter_dados_tabela(url):
    requisicao = requests.get(url)
    html = requisicao.text
    soup = BeautifulSoup(html, "html.parser")

    tabelas = soup.find_all("table", class_="wikitable")
    tabela_classificacao = None
    dados = []

    # Localiza a tabela de classificação na página
    for tabela in tabelas:
        if "classificação" in tabela.get_text().lower():
            tabela_classificacao = tabela
            break

    if tabela_classificacao:
        linhas = tabela_classificacao.find_all("tr")[1:]
        for linha in linhas:
            colunas = linha.find_all("td")
            if len(colunas) >= 11:
                posicao = colunas[0].text.strip()
                time = colunas[1].text.strip()
                pontuacao = colunas[2].text.strip()
                aproveitamento = colunas[10].text.strip()

                dados.append([time, pontuacao, posicao, aproveitamento])

    return dados

# URL da página da tabela (pode ser qualquer tabela HTML)
url_tabela = "https://exemplo.com/tabela"

# Obtém os dados da tabela
dados_tabela = obter_dados_tabela(url_tabela)

# Arquivo Excel
arquivo_excel = "Caminho/Para/Seu/Arquivo.xlsx"
planilha_desejada = "Nome da Planilha"
colunas = ["A", "B", "C", "D"]  

# Função para inserir os dados no Excel
def inserir_dados_excel(arquivo, planilha, colunas, dados):
    workbook = load_workbook(arquivo)
    sheet = workbook[planilha]

    for i, coluna in enumerate(colunas):
        valores = [dados[j][i] for j in range(len(dados))]
        for j, valor in enumerate(valores):
            linha = j + 2
            if coluna == "D":  
                valor = f"{valor}%"
            sheet[f"{coluna}{linha}"] = valor

    workbook.save(arquivo)

# Função para visualizar as colunas após a inserção
def visualizar_colunas(arquivo, planilha, colunas):
    workbook = load_workbook(arquivo)
    sheet = workbook[planilha]

    for coluna in colunas:
        valores_coluna = [cell.value for cell in sheet[coluna][2:len(sheet[coluna]) + 1]]
        print(f"Coluna {coluna}:")
        for valor in valores_coluna:
            print(valor)
        print()

# Inserir os dados no Excel
try:
    inserir_dados_excel(arquivo_excel, planilha_desejada, colunas, dados_tabela)
    print("Inserção de dados bem-sucedida!")
except Exception as e:
    print("Ocorreu um erro durante a inserção de dados:", str(e))

# Visualizar as colunas após a inserção
print("Colunas após a inserção:")
visualizar_colunas(arquivo_excel, planilha_desejada, colunas)


